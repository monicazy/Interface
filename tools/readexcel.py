# -*- coding: utf-8 -*-
# @Time  : 2018/11/8 23:12
# @Author : Monica
# @Email : 498194410@qq.com
# @File  : readexcel.py
from openpyxl import load_workbook
from tools.do_config import Readconfig
from tools import project_path
import pandas as pd
from tools.get_data import GetData
from tools.do_regx import DoRegx

class ReadExcel:
    def __init__(self):
        self.file_name = project_path.test_case_path
        # 打开excel
        self.wb = load_workbook(self.file_name)

    def read_excel(self):
        sheet_name_sub = eval(Readconfig().read_config(project_path.case_config_path, "DATA", "data"))  # 得到一个字典
        sheet_name = [key for key in sorted(sheet_name_sub)]  # 获取字典里存放的keys转成list(即获取sheet_name)

        data = []
        for n in range(len(sheet_name)):  # 遍历sheet索引
            df = pd.read_excel(self.file_name, sheet_name[n])
            for i in df.index.values:  # 获取行号的索引，并对其进行遍历：
                if sheet_name_sub[sheet_name[n]] == "all":  # 如果value等于all就全部遍历添加
                    row_data = df.loc[df.index[i],
                                      ['id', "module", "title", "url", "data", "check_sql", "method", "ExpectedResult"]].to_dict()
                    row_data["sheet_name"] = sheet_name[n]
                    # 替换变量值
                    row_data["data"] = DoRegx().do_regx(row_data["data"])
                    if eval(row_data["check_sql"]) != None:
                        row_data["check_sql"] = DoRegx().do_regx(row_data["check_sql"])
                    data.append(row_data)
                else:
                    if i in sheet_name_sub[sheet_name[n]]:  # 如果i在字典中对应的value值（即指定用例索引值）里存在就添加
                        row_data = df.loc[df.index[i],
                                          ['id', "module", "title", "url", "data", "check_sql", "method",
                                           "ExpectedResult"]].to_dict()
                        row_data["sheet_name"] = sheet_name[n]
                        # 替换变量值
                        row_data["data"] = DoRegx().do_regx(row_data["data"])
                        if eval(row_data["check_sql"]) != None:
                            row_data["check_sql"] = DoRegx().do_regx(row_data["check_sql"])
                        data.append(row_data)
            ReadExcel().update_excel(int(getattr(GetData, "tel")) + 1)  # 更新变量表
        return data

    def write_back(self, sheet_name, row, col,result):
        sheet = self.wb[sheet_name]
        sheet.cell(row=row, column=col).value = result
        self.wb.save(self.file_name)

    def update_excel(self, var_value):
        sheet = self.wb["init"]
        sheet.cell(row=2, column=1).value = var_value
        self.wb.save(self.file_name)

    @staticmethod
    def replace_var(str_name, var_name, rep_name):
        if str_name.find(var_name) != -1:
            str_name = str_name.replace(var_name, rep_name)
        else:
            str_name = str_name
        return str_name


if __name__ == '__main__':
    print(ReadExcel().read_excel())
